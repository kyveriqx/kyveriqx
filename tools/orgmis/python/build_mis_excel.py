"""
Build a board-ready MIS workbook (multi-sheet, formatted).
Reads .tmp/financial_summary.json and writes MIS_FY24-25.xlsx
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference, BarChart3D
from openpyxl.chart.label import DataLabelList
import json, os

from config_loader import BRANDING, WORKDIR, hex_no_hash

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TMP = WORKDIR
OUT = os.path.join(WORKDIR, 'MIS.xlsx')

with open(os.path.join(TMP, 'financial_summary.json')) as f:
    s = json.load(f)

# ----- Styles -----
NAVY = hex_no_hash(BRANDING.get('primaryColor'), '1F3864')
GOLD = hex_no_hash(BRANDING.get('accentColor'), 'BF8F00')
LIGHT_NAVY = 'D9E1F2'
LIGHT_GOLD = 'FFF2CC'
GREEN = '548235'
RED = 'C00000'
BORDER_THIN = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF'),
)
BORDER_HEAVY_BOTTOM = Border(bottom=Side(style='medium', color='1F3864'))

def title_style(cell, text, size=16, color=NAVY):
    cell.value = text
    cell.font = Font(name='Calibri', size=size, bold=True, color=color)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def header_style(cell, text):
    cell.value = text
    cell.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor=NAVY)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER_THIN

def value_cell(cell, val, fmt='#,##0', bold=False, fill=None, color=None, italic=False):
    cell.value = val
    cell.number_format = fmt
    f = Font(name='Calibri', size=11, bold=bold, italic=italic, color=color or '000000')
    cell.font = f
    cell.alignment = Alignment(horizontal='right')
    cell.border = BORDER_THIN
    if fill:
        cell.fill = PatternFill('solid', fgColor=fill)

def label_cell(cell, val, bold=False, indent=0, fill=None, italic=False, color=None):
    cell.value = val
    cell.font = Font(name='Calibri', size=11, bold=bold, italic=italic, color=color or '000000')
    cell.alignment = Alignment(horizontal='left', indent=indent)
    cell.border = BORDER_THIN
    if fill:
        cell.fill = PatternFill('solid', fgColor=fill)

def to_cr(v):
    return v / 1e7  # crore

# ============================
wb = openpyxl.Workbook()

# ====== Sheet 1: Cover ======
ws = wb.active
ws.title = 'Cover'
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 90

# Brand bar
for c in range(1, 11):
    ws.cell(2, c).fill = PatternFill('solid', fgColor=NAVY)
ws.row_dimensions[2].height = 6

PERIOD = BRANDING.get('reportingPeriod', 'FY 2024-25')

ws.cell(4, 2).value = BRANDING.get('companyName', 'Your Company')
ws.cell(4, 2).font = Font(name='Calibri', size=28, bold=True, color=NAVY)
ws.cell(6, 2).value = BRANDING.get('tagline', 'Board Report')
ws.cell(6, 2).font = Font(name='Calibri', size=14, italic=True, color='595959')

ws.cell(10, 2).value = "MANAGEMENT INFORMATION SYSTEM (MIS)"
ws.cell(10, 2).font = Font(name='Calibri', size=20, bold=True, color=GOLD)
ws.cell(12, 2).value = f"Financial Performance Review — {BRANDING.get('reportingPeriod', 'FY 2024-25')}"
ws.cell(12, 2).font = Font(name='Calibri', size=16, bold=True, color=NAVY)
ws.cell(13, 2).value = s['period']
ws.cell(13, 2).font = Font(name='Calibri', size=11, italic=True, color='595959')

# TOC
ws.cell(17, 2).value = "Contents"
ws.cell(17, 2).font = Font(name='Calibri', size=14, bold=True, color=NAVY)
toc = [
    ("1. Executive Summary", "Executive Summary"),
    (f"2. Profit & Loss Statement ({PERIOD})", "P&L Statement"),
    ("3. Monthly Revenue & EBITDA Trend", "Monthly Trend"),
    ("4. Top Customers", "Top Customers"),
    ("5. Top Vendors", "Top Vendors"),
    ("6. Geography & Currency Mix", "Geography"),
    ("7. Working Capital Snapshot", "Working Capital"),
    ("8. GL Category Summary", "GL Summary"),
    ("9. Notes & Assumptions", "Notes"),
]
for i, (label, _) in enumerate(toc):
    ws.cell(18+i, 2).value = label
    ws.cell(18+i, 2).font = Font(name='Calibri', size=11, color=NAVY)

ws.cell(30, 2).value = f"Prepared for: {BRANDING.get('preparedFor', 'Board of Directors')}"
ws.cell(30, 2).font = Font(name='Calibri', size=10, italic=True, color='595959')
ws.cell(31, 2).value = "Currency: INR (figures in Crores unless stated)"
ws.cell(31, 2).font = Font(name='Calibri', size=10, italic=True, color='595959')

# ====== Sheet 2: Executive Summary ======
ws = wb.create_sheet('Executive Summary')
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 42
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 18

title_style(ws.cell(2, 2), f"Executive Summary — {PERIOD}", size=18)
ws.cell(3, 2).value = s['period']
ws.cell(3, 2).font = Font(size=10, italic=True, color='595959')

pl = s['pl_actual']
pla = s['pl_annualized']

# KPI tiles - 6 across
kpis = [
    ('Revenue', pl['Revenue'], pla['Revenue'], None),
    ('Gross Profit', pl['Gross Profit'], pla['Gross Profit'], pl['Gross Margin %']),
    ('EBITDA', pl['EBITDA'], pla['EBITDA'], pl['EBITDA Margin %']),
    ('EBIT', pl['EBIT'], pla['EBIT'], pl['EBIT Margin %']),
    ('PBT', pl['PBT'], pla['PBT'], pl['PBT Margin %']),
    ('PAT (Net Profit)', pl['PAT'], pla['PAT'], pl['PAT / Net Margin %']),
]

# Header row for KPI table
r = 6
header_style(ws.cell(r, 2), 'Key Metric')
header_style(ws.cell(r, 3), '11M Actual\n(₹ Cr)')
header_style(ws.cell(r, 4), 'Annualized\n(₹ Cr)')
header_style(ws.cell(r, 5), 'Margin %')
ws.row_dimensions[r].height = 35
r += 1
for label, act, ann, margin in kpis:
    is_neg = act < 0
    color = RED if is_neg else None
    label_cell(ws.cell(r, 2), label, bold=True, fill=LIGHT_NAVY)
    value_cell(ws.cell(r, 3), to_cr(act), fmt='#,##0.00', bold=True, color=color)
    value_cell(ws.cell(r, 4), to_cr(ann), fmt='#,##0.00', color=color)
    if margin is not None:
        value_cell(ws.cell(r, 5), margin/100, fmt='0.00%', color=color)
    else:
        value_cell(ws.cell(r, 5), '', fmt='@')
    r += 1

# Commentary block
r += 2
title_style(ws.cell(r, 2), 'Key Observations', size=14)
r += 1
obs = [
    f"• Revenue stood at ₹{to_cr(pl['Revenue']):.2f} Cr for 11 months (annualized ₹{to_cr(pla['Revenue']):.2f} Cr).",
    f"• Gross margin healthy at {pl['Gross Margin %']:.1f}% — pricing and product mix remain strong.",
    f"• EBITDA of ₹{to_cr(pl['EBITDA']):.2f} Cr ({pl['EBITDA Margin %']:.1f}% margin) — operating cost base requires optimization.",
    f"• Depreciation of ₹{to_cr(pl['Depreciation']):.2f} Cr drove EBIT into negative territory.",
    f"• {'Marginal loss' if pl['PAT'] < 0 else 'Net profit'} at PAT level of ₹{to_cr(abs(pl['PAT'])):.2f} Cr.",
    f"• Strong upward monthly revenue trajectory (Apr'24 ₹{to_cr(list(s['monthly_revenue'].values())[0]):.2f} Cr → Jan'25 ₹{to_cr(sorted(s['monthly_revenue'].items())[-2][1]):.2f} Cr).",
    f"• Working capital: Open AR ₹{to_cr(s['total_ar_open']):.2f} Cr vs Open AP ₹{to_cr(s['total_ap_open']):.2f} Cr.",
]
for o in obs:
    ws.cell(r, 2).value = o
    ws.cell(r, 2).font = Font(size=11)
    ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical='top')
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    ws.row_dimensions[r].height = 22
    r += 1

# ====== Sheet 3: P&L Statement ======
ws = wb.create_sheet('P&L Statement')
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 20
ws.column_dimensions['E'].width = 14

title_style(ws.cell(2, 2), f'Profit & Loss Statement — {PERIOD}', size=18)
ws.cell(3, 2).value = f"All figures in INR Crores. {s['period']}"
ws.cell(3, 2).font = Font(size=10, italic=True, color='595959')

r = 5
header_style(ws.cell(r, 2), 'Line Item')
header_style(ws.cell(r, 3), '11M Actual')
header_style(ws.cell(r, 4), 'Annualized')
header_style(ws.cell(r, 5), '% of Rev')
ws.row_dimensions[r].height = 28
r += 1

def pl_row(label, key, indent=0, bold=False, subtotal=False, italic=False, is_margin_key=None):
    global r
    fill = LIGHT_NAVY if subtotal else None
    label_cell(ws.cell(r, 2), label, bold=bold, indent=indent, fill=fill, italic=italic)
    v_act = pl[key]; v_ann = pla[key]
    color = RED if v_act < 0 else None
    value_cell(ws.cell(r, 3), to_cr(v_act), fmt='#,##0.00;(#,##0.00)', bold=bold, fill=fill, color=color)
    value_cell(ws.cell(r, 4), to_cr(v_ann), fmt='#,##0.00;(#,##0.00)', bold=bold, fill=fill, color=color)
    if is_margin_key and pl['Revenue']:
        pct = v_act / pl['Revenue']
        value_cell(ws.cell(r, 5), pct, fmt='0.0%', bold=bold, fill=fill, color=color)
    else:
        value_cell(ws.cell(r, 5), '', fmt='@', bold=bold, fill=fill)
    r += 1

pl_row('Revenue from Operations', 'Revenue', bold=True, subtotal=True, is_margin_key=True)
pl_row('Less: Cost of Goods Sold (COGS)', 'COGS', indent=1, is_margin_key=True)
pl_row('Gross Profit', 'Gross Profit', bold=True, subtotal=True, is_margin_key=True)
pl_row('Less: Operating Expenses', 'Operating Expenses', indent=1, is_margin_key=True)
pl_row('Add: Other Income', 'Other Income', indent=1, is_margin_key=True)
pl_row('Less: Other Expenses', 'Other Expenses', indent=1, is_margin_key=True)
pl_row('EBITDA', 'EBITDA', bold=True, subtotal=True, is_margin_key=True)
pl_row('Less: Depreciation & Amortization', 'Depreciation', indent=1, is_margin_key=True)
pl_row('EBIT (Operating Profit)', 'EBIT', bold=True, subtotal=True, is_margin_key=True)
pl_row('Less: Finance Costs', 'Finance Costs', indent=1, is_margin_key=True)
pl_row('Profit Before Tax (PBT)', 'PBT', bold=True, subtotal=True, is_margin_key=True)
pl_row('Less: Tax (Indicative @25.17%)', 'Tax (Indicative @25.17%)', indent=1, italic=True, is_margin_key=True)
pl_row('Profit After Tax (PAT) — Net Profit', 'PAT', bold=True, subtotal=True, is_margin_key=True)

# Margin summary block
r += 2
title_style(ws.cell(r, 2), 'Margin Profile', size=14)
r += 1
margin_rows = [
    ('Gross Margin', 'Gross Margin %'),
    ('EBITDA Margin', 'EBITDA Margin %'),
    ('EBIT Margin', 'EBIT Margin %'),
    ('PBT Margin', 'PBT Margin %'),
    ('PAT / Net Margin', 'PAT / Net Margin %'),
]
header_style(ws.cell(r, 2), 'Margin')
header_style(ws.cell(r, 3), 'Value')
r += 1
for lbl, k in margin_rows:
    label_cell(ws.cell(r, 2), lbl)
    is_neg = pl[k] < 0
    value_cell(ws.cell(r, 3), pl[k]/100, fmt='0.00%', color=RED if is_neg else GREEN, bold=True)
    r += 1

# ====== Sheet 4: Monthly Trend ======
ws = wb.create_sheet('Monthly Trend')
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
for c in range(2, 8):
    ws.column_dimensions[get_column_letter(c)].width = 16

title_style(ws.cell(2, 2), f'Monthly Revenue Trend — {PERIOD}', size=18)

r = 5
header_style(ws.cell(r, 2), 'Month')
header_style(ws.cell(r, 3), 'Revenue (₹ Cr)')
ws.row_dimensions[r].height = 28

month_labels = {'2024-04':"Apr-24",'2024-05':"May-24",'2024-06':"Jun-24",'2024-07':"Jul-24",
                '2024-08':"Aug-24",'2024-09':"Sep-24",'2024-10':"Oct-24",'2024-11':"Nov-24",
                '2024-12':"Dec-24",'2025-01':"Jan-25",'2025-02':"Feb-25"}

r0 = 6
mr_items = sorted(s['monthly_revenue'].items())
for i, (ym, rev) in enumerate(mr_items):
    label_cell(ws.cell(r0+i, 2), month_labels.get(ym, ym))
    value_cell(ws.cell(r0+i, 3), to_cr(rev), fmt='#,##0.00', bold=True)

# total row
tr = r0 + len(mr_items)
label_cell(ws.cell(tr, 2), 'Total (11M)', bold=True, fill=LIGHT_NAVY)
value_cell(ws.cell(tr, 3), to_cr(sum(s['monthly_revenue'].values())), fmt='#,##0.00', bold=True, fill=LIGHT_NAVY)

# Chart
chart = LineChart()
chart.title = "Monthly Revenue Trend (₹ Cr)"
chart.style = 12
chart.x_axis.title = "Month"
chart.y_axis.title = "Revenue (₹ Cr)"
data = Reference(ws, min_col=3, min_row=5, max_row=r0+len(mr_items)-1, max_col=3)
cats = Reference(ws, min_col=2, min_row=r0, max_row=r0+len(mr_items)-1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.height = 10
chart.width = 22
chart.dataLabels = DataLabelList(showVal=True)
ws.add_chart(chart, "E5")

# ====== Sheet 5: Top Customers ======
ws = wb.create_sheet('Top Customers')
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 5
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 22

title_style(ws.cell(2, 2), 'Top Customers by Invoice Value', size=18)
ws.cell(3, 2).value = "Source: Sales Invoice Lines (recent period sample). Currency: as billed (mixed)."
ws.cell(3, 2).font = Font(size=10, italic=True, color='595959')

r = 5
header_style(ws.cell(r, 2), '#')
header_style(ws.cell(r, 3), 'Customer')
header_style(ws.cell(r, 4), 'Invoice Value')
ws.row_dimensions[r].height = 28
r += 1

_top_cust = [(n, a) for n, a in s.get('top_customers') or [] if a]
if _top_cust:
    for i, (name, amt) in enumerate(_top_cust[:10], 1):
        value_cell(ws.cell(r, 2), i, fmt='0')
        label_cell(ws.cell(r, 3), name)
        value_cell(ws.cell(r, 4), amt, fmt='#,##0')
        r += 1
    # Pie chart for customer concentration (only if we have real data).
    chart = PieChart()
    chart.title = "Top 10 Customer Concentration"
    labels = Reference(ws, min_col=3, min_row=6, max_row=5+len(_top_cust[:10]))
    data = Reference(ws, min_col=4, min_row=5, max_row=5+len(_top_cust[:10]))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.height = 10
    chart.width = 16
    ws.add_chart(chart, "F5")
else:
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    label_cell(
        ws.cell(r, 2),
        "Insufficient sales data to compute top customers — re-upload "
        "with customer name + amount columns.",
        italic=True, color='595959',
    )
    ws.row_dimensions[r].height = 36

# ====== Sheet 6: Top Vendors ======
ws = wb.create_sheet('Top Vendors')
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 5
ws.column_dimensions['C'].width = 40
ws.column_dimensions['D'].width = 22

title_style(ws.cell(2, 2), 'Top Vendors by Purchase Value', size=18)
ws.cell(3, 2).value = "Source: Purchase Invoice Lines. Currency: as billed (mixed)."
ws.cell(3, 2).font = Font(size=10, italic=True, color='595959')

r = 5
header_style(ws.cell(r, 2), '#')
header_style(ws.cell(r, 3), 'Vendor')
header_style(ws.cell(r, 4), 'Purchase Value')
ws.row_dimensions[r].height = 28
r += 1

_top_vend = [(n, a) for n, a in s.get('top_vendors') or [] if a]
if _top_vend:
    for i, (name, amt) in enumerate(_top_vend[:10], 1):
        value_cell(ws.cell(r, 2), i, fmt='0')
        label_cell(ws.cell(r, 3), name)
        value_cell(ws.cell(r, 4), amt, fmt='#,##0')
        r += 1
else:
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    label_cell(
        ws.cell(r, 2),
        "Insufficient purchase data to compute top vendors — re-upload "
        "with vendor name + amount columns.",
        italic=True, color='595959',
    )
    ws.row_dimensions[r].height = 36

# ====== Sheet 7: Geography ======
ws = wb.create_sheet('Geography')
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
for c in range(2, 8):
    ws.column_dimensions[get_column_letter(c)].width = 18

title_style(ws.cell(2, 2), 'Revenue by Geography & Currency', size=18)

r = 5
header_style(ws.cell(r, 2), 'Country')
header_style(ws.cell(r, 3), 'Invoice Value')
ws.row_dimensions[r].height = 28
r += 1
country_sorted = sorted(s['country_rev'].items(), key=lambda x: -x[1])
r_start = r
for code, amt in country_sorted:
    label_cell(ws.cell(r, 2), code or 'Unspecified')
    value_cell(ws.cell(r, 3), amt, fmt='#,##0')
    r += 1

# Currency mix to right
r2 = 5
header_style(ws.cell(r2, 5), 'Currency')
header_style(ws.cell(r2, 6), 'Invoice Value')
ws.row_dimensions[r2].height = 28
r2 += 1
curr_sorted = sorted(s['currency_rev'].items(), key=lambda x: -x[1])
for code, amt in curr_sorted:
    label_cell(ws.cell(r2, 5), code or 'INR (Local)')
    value_cell(ws.cell(r2, 6), amt, fmt='#,##0')
    r2 += 1

# Country pie
chart = PieChart()
chart.title = "Revenue by Country"
labels = Reference(ws, min_col=2, min_row=6, max_row=r-1)
data = Reference(ws, min_col=3, min_row=5, max_row=r-1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)
chart.height = 9; chart.width = 14
ws.add_chart(chart, "H5")

# Currency pie
chart2 = PieChart()
chart2.title = "Revenue by Currency"
labels = Reference(ws, min_col=5, min_row=6, max_row=r2-1)
data = Reference(ws, min_col=6, min_row=5, max_row=r2-1)
chart2.add_data(data, titles_from_data=True)
chart2.set_categories(labels)
chart2.height = 9; chart2.width = 14
ws.add_chart(chart2, "H22")

# ====== Sheet 8: Working Capital ======
ws = wb.create_sheet('Working Capital')
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 38
ws.column_dimensions['C'].width = 20

title_style(ws.cell(2, 2), 'Working Capital Snapshot', size=18)
ws.cell(3, 2).value = "Closing balances from GL Entry (₹ Cr)"
ws.cell(3, 2).font = Font(size=10, italic=True, color='595959')

r = 5
header_style(ws.cell(r, 2), 'Item')
header_style(ws.cell(r, 3), 'Value (₹ Cr)')
ws.row_dimensions[r].height = 28
r += 1
items = [
    ('Trade Receivables (Open AR)', s['total_ar_open']),
    ('Trade Payables (Open AP)', s['total_ap_open']),
    ('Net Working Capital (AR - AP)', s['total_ar_open'] - s['total_ap_open']),
    ('AR Days (on annualized revenue)', (s['total_ar_open'] / pla['Revenue']) * 365 if pla['Revenue'] else 0),
    ('AP Days (on annualized COGS)', (s['total_ap_open'] / pla['COGS']) * 365 if pla['COGS'] else 0),
]
for lbl, v in items:
    label_cell(ws.cell(r, 2), lbl)
    if 'Days' in lbl:
        value_cell(ws.cell(r, 3), v, fmt='0.0', bold=True)
    else:
        value_cell(ws.cell(r, 3), to_cr(v), fmt='#,##0.00', bold=True)
    r += 1

# ====== Sheet 9: GL Category Summary ======
ws = wb.create_sheet('GL Summary')
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 35
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 22

title_style(ws.cell(2, 2), 'GL Activity by Category', size=18)

r = 5
header_style(ws.cell(r, 2), 'Category')
header_style(ws.cell(r, 3), 'FY 24-25 (₹ Cr)')
header_style(ws.cell(r, 4), 'FY 23-24 (₹ Cr) Partial')
ws.row_dimensions[r].height = 28
r += 1

cats = s['gl_account_summary'].get('FY24-25', {})
cats_prev = s['gl_account_summary'].get('FY23-24', {})
all_cats = sorted(set(list(cats.keys()) + list(cats_prev.keys())))
for c in all_cats:
    label_cell(ws.cell(r, 2), c)
    value_cell(ws.cell(r, 3), to_cr(cats.get(c, 0)), fmt='#,##0.00;(#,##0.00)')
    value_cell(ws.cell(r, 4), to_cr(cats_prev.get(c, 0)), fmt='#,##0.00;(#,##0.00)')
    r += 1

# ====== Sheet 10: Notes ======
ws = wb.create_sheet('Notes')
ws.sheet_view.showGridLines = False
ws.column_dimensions['A'].width = 3
ws.column_dimensions['B'].width = 110

title_style(ws.cell(2, 2), 'Notes & Assumptions', size=18)
notes = [
    "1. Reporting Period: FY 2024-25 covers Apr 2024 to Feb 2025 (11 months of actual postings).",
    "   Data ends Feb 2025 with only stub activity; March 2025 figures are not yet available.",
    "2. Annualization: 11-month actuals scaled by 12/11 to project full-year run rate.",
    "3. Income Tax: An indicative provision at 25.17% (standard Indian corporate rate including surcharge & cess) is applied to positive PBT only.",
    "   No tax expense was directly recorded in the GL Entry exports.",
    "4. Account classification follows NAV/Business Central chart-of-accounts conventions:",
    "   6xxx = Revenue, 7xxx = COGS, 8000-8799 = OpEx, 8800-8899 = Depreciation, 9xxx = Other/Finance.",
    "5. Top Customers and Top Vendors are based on Sales/Purchase Invoice Line data, which covers a sample period (Nov 2024 - Feb 2025).",
    "   Full revenue from GL ledger (₹42.0 Cr) is significantly larger than the sample invoice totals.",
    "6. AR/AP balances are computed from GL closing balances of receivables (2310-2330) and payables (5410-5420) accounts.",
    "7. Currency: Local currency (LCY) treated as INR. Foreign currency invoices are presented in their as-billed values for the Customer/Vendor breakdowns.",
    "8. All figures shown in ₹ Crores unless explicitly labelled.",
    "9. Prior year (FY 23-24) data is partial (Dec 2023 - Mar 2024 only); direct YoY comparison is not reliable.",
]
r = 4
for n in notes:
    ws.cell(r, 2).value = n
    ws.cell(r, 2).font = Font(size=11)
    ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical='top')
    ws.row_dimensions[r].height = 30
    r += 1

# Save
wb.save(OUT)
print(f"\nMIS workbook saved: {OUT}")
print(f"Sheets: {wb.sheetnames}")
