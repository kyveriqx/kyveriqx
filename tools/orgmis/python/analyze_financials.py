"""
Analyze GL Entry + related ledgers to produce a board-ready financial summary.
Outputs: <WORKDIR>/financial_summary.json
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict
from datetime import datetime, date
import json
import os

from config_loader import BRANDING, OUTLOOK, FILES, WORKDIR, all_files

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TMP = WORKDIR
os.makedirs(TMP, exist_ok=True)

# ---------- Helpers ----------
def load_sheet(path, header_row=3):
    # Accept absolute path; if relative, anchor on project base
    p = path if os.path.isabs(path) else os.path.join(BASE, path)
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb.active
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column+1)]
    rows = []
    for r in range(header_row+1, ws.max_row+1):
        rec = {}
        empty = True
        for c, h in enumerate(headers, 1):
            v = ws.cell(r, c).value
            rec[h] = v
            if v not in (None, ''):
                empty = False
        if not empty:
            rows.append(rec)
    return rows

def to_date(v):
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    if isinstance(v, str):
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
            try: return datetime.strptime(v, fmt).date()
            except: pass
    return None

def fy_of(d):
    """Indian FY: Apr-Mar. Apr2024-Mar2025 = FY24-25"""
    if d is None: return None
    if d.month >= 4: return f"FY{str(d.year)[-2:]}-{str(d.year+1)[-2:]}"
    return f"FY{str(d.year-1)[-2:]}-{str(d.year)[-2:]}"

# ---------- Load GL (multiple files supported) ----------
print("Loading GL Entry...")
gl = []
for p in all_files('gl'):
    rows = load_sheet(p)
    if not rows:
        # Try header on row 1 as fallback
        rows = load_sheet(p, header_row=1)
    print(f"  + {os.path.basename(p)}: {len(rows)} rows")
    gl.extend(rows)
print(f"  Total GL rows: {len(gl)}")

# ---------- Classify accounts ----------
# Based on NAV/BC chart of accounts conventions
def classify(acct):
    a = int(acct) if str(acct).isdigit() else 0
    if 1000 <= a <= 1999: return 'Fixed Assets'
    if 2000 <= a <= 2199: return 'Inventory'
    if 2200 <= a <= 2399: return 'Receivables'
    if 2400 <= a <= 2799: return 'Other Current Assets'
    if 2800 <= a <= 2999: return 'Cash & Bank'
    if 3000 <= a <= 3999: return 'Equity'
    if 4000 <= a <= 4999: return 'Long-Term Liabilities'
    if 5000 <= a <= 5499: return 'Trade Payables'
    if 5500 <= a <= 5799: return 'Tax Payables (VAT/GST)'
    if 5800 <= a <= 5999: return 'Statutory Payables'
    if 6000 <= a <= 6999: return 'Revenue'
    if 7000 <= a <= 7999: return 'COGS'
    if 8000 <= a <= 8799: return 'Operating Expenses'
    if 8800 <= a <= 8899: return 'Depreciation'
    if 8900 <= a <= 8999: return 'Other OpEx'
    if 9000 <= a <= 9199: return 'Other Income'
    if 9200 <= a <= 9499: return 'Other Expenses'
    if 9500 <= a <= 9999: return 'Finance Costs'
    return 'Unclassified'

# ---------- Period filter ----------
FY24_25_START = date(2024, 4, 1)
FY24_25_END   = date(2025, 3, 31)
FY23_24_START = date(2023, 4, 1)
FY23_24_END   = date(2024, 3, 31)

# Aggregate by category and FY
by_fy_cat = defaultdict(lambda: defaultdict(float))     # fy -> cat -> amt
by_fy_acct = defaultdict(lambda: defaultdict(float))    # fy -> acct -> amt
by_month_cat = defaultdict(lambda: defaultdict(float))  # YYYY-MM -> cat -> amt
acct_desc = {}
latest_posting_date = None

for r in gl:
    d = to_date(r['Posting Date'])
    if d is None: continue
    if latest_posting_date is None or d > latest_posting_date:
        latest_posting_date = d
    acct = r['G/L Account No.']
    amt = float(r['Amount'] or 0)
    cat = classify(acct)
    fy = fy_of(d)
    by_fy_cat[fy][cat] += amt
    by_fy_acct[fy][acct] += amt
    by_month_cat[d.strftime('%Y-%m')][cat] += amt
    if acct not in acct_desc and r.get('Description'):
        acct_desc[acct] = str(r['Description'])[:60]

# ---------- Build P&L for FY24-25 ----------
def build_pl(fy):
    cats = by_fy_cat[fy]
    # Revenue posted as credits (negative). Flip sign.
    revenue = -cats.get('Revenue', 0)
    cogs = cats.get('COGS', 0)
    gross_profit = revenue - cogs

    opex_admin = cats.get('Operating Expenses', 0) + cats.get('Other OpEx', 0)
    depreciation = cats.get('Depreciation', 0)
    other_income = -cats.get('Other Income', 0)  # credits flipped
    other_exp = cats.get('Other Expenses', 0)
    finance_cost = cats.get('Finance Costs', 0)

    ebitda = gross_profit - opex_admin + other_income - other_exp
    ebit = ebitda - depreciation
    pbt = ebit - finance_cost
    tax = max(0, pbt) * 0.2517  # indicative Indian corp tax
    pat = pbt - tax

    return {
        'Revenue': revenue,
        'COGS': cogs,
        'Gross Profit': gross_profit,
        'Gross Margin %': (gross_profit/revenue*100) if revenue else 0,
        'Operating Expenses': opex_admin,
        'Other Income': other_income,
        'Other Expenses': other_exp,
        'EBITDA': ebitda,
        'EBITDA Margin %': (ebitda/revenue*100) if revenue else 0,
        'Depreciation': depreciation,
        'EBIT': ebit,
        'EBIT Margin %': (ebit/revenue*100) if revenue else 0,
        'Finance Costs': finance_cost,
        'PBT': pbt,
        'PBT Margin %': (pbt/revenue*100) if revenue else 0,
        'Tax (Indicative @25.17%)': tax,
        'PAT': pat,
        'PAT / Net Margin %': (pat/revenue*100) if revenue else 0,
    }

pl_24_25 = build_pl('FY24-25')
pl_23_24 = build_pl('FY23-24')

# Annualize FY24-25 from 11 months -> 12 months for projection
months_in_fy = 11
ann_factor = 12 / months_in_fy
pl_24_25_ann = {k: (v * ann_factor if '%' not in k else v) for k,v in pl_24_25.items()}

print("\n=== FY24-25 P&L (11 months actuals, Apr'24-Feb'25) ===")
for k, v in pl_24_25.items():
    suffix = '%' if '%' in k else ''
    print(f"  {k:<28}: {v:>18,.2f}{suffix}")

print("\n=== FY24-25 P&L (Annualized) ===")
for k, v in pl_24_25_ann.items():
    suffix = '%' if '%' in k else ''
    print(f"  {k:<28}: {v:>18,.2f}{suffix}")

print("\n=== FY23-24 P&L (Q4 only, Dec'23-Mar'24 ~ 4 months) ===")
for k, v in pl_23_24.items():
    suffix = '%' if '%' in k else ''
    print(f"  {k:<28}: {v:>18,.2f}{suffix}")

# ---------- Operational KPIs ----------
print("\nLoading operational data...")
# Load sales invoice header (use openpyxl directly because we need a specific sheet)
def load_named_sheet(path, sheet, header_row=3):
    p = path if os.path.isabs(path) else os.path.join(BASE, path)
    wb = openpyxl.load_workbook(p, data_only=True)
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column+1)]
    rows = []
    for r in range(header_row+1, ws.max_row+1):
        rec = {}
        empty = True
        for c, h in enumerate(headers, 1):
            v = ws.cell(r, c).value
            rec[h] = v
            if v not in (None, ''): empty = False
        if not empty: rows.append(rec)
    return rows

# Fuzzy column lookup — users upload exports with very different header
# conventions (Business Central, Tally, Zoho, Busy, SAP, custom GST registers).
# Matching strategy:
#   1. exact key match
#   2. case-/punctuation-insensitive exact match
#   3. token-set match — every alphanumeric token in the candidate must appear
#      as a whole token in the header. So "Customer" matches "Customer Name"
#      but NOT "Customer GST No." (the GST/No tokens make it a strictly
#      different concept). Among matches we pick the header whose token-set
#      has the smallest extras count → most-specific wins.
import re as _re_cols

def _norm_col(k):
    if k is None:
        return ""
    return _re_cols.sub(r"[^a-z0-9]+", "", str(k).lower())

def _tokens(k):
    if k is None:
        return frozenset()
    return frozenset(t for t in _re_cols.split(r"[^a-z0-9]+", str(k).lower()) if t)

def find_value(row, candidates):
    if not row:
        return None
    for cand in candidates:
        if cand in row and row[cand] not in (None, ""):
            return row[cand]
    norm_row = {_norm_col(k): (k, v) for k, v in row.items()}
    for cand in candidates:
        nk = _norm_col(cand)
        if nk in norm_row and norm_row[nk][1] not in (None, ""):
            return norm_row[nk][1]
    # Token-set match: candidate tokens must be a subset of header tokens.
    # Prefer the header with the fewest extra tokens (most specific).
    tokenized_row = [(_tokens(k), v) for k, v in row.items() if v not in (None, "")]
    best = None
    best_extras = None
    for cand in candidates:
        ct = _tokens(cand)
        if not ct:
            continue
        for ht, v in tokenized_row:
            if ct.issubset(ht):
                extras = len(ht - ct)
                if best_extras is None or extras < best_extras:
                    best = v
                    best_extras = extras
        if best is not None:
            return best
    return None

def list_headers(rows, limit=2):
    """Collect a representative set of column headers from a list of rows.
       Used by diagnostic logging when top-customer / top-vendor extraction
       produces empty results — the trace shows me exactly what columns the
       user's file has so I can extend the candidate list next iteration."""
    seen = []
    for r in rows[:limit] if rows else []:
        for k in (r.keys() if hasattr(r, "keys") else []):
            if k not in seen:
                seen.append(k)
    return seen

# Column name candidates by field (most-specific first). Expanded to cover
# Tally, Zoho, Busy, BC, SAP, and common flat-register exports.
DOC_NO_KEYS    = ['Document No.', 'Document No', 'Invoice No.', 'Invoice No',
                  'No.', 'No', 'Doc No', 'Voucher No.', 'Voucher No',
                  'Voucher', 'Voucher Number', 'Bill No', 'Bill No.',
                  'Bill Number', 'Reference No', 'Ref No', 'Transaction Id',
                  'Transaction No']
AMOUNT_KEYS    = ['Line Amount', 'Amount', 'Invoice Amount', 'Total Amount',
                  'Net Amount', 'Amount (LCY)', 'Taxable Value', 'Value',
                  'Net', 'Total', 'Net Value', 'Total Value', 'Bill Amount',
                  'Sales Value', 'Net Sales', 'Total Sales', 'Grand Total',
                  'Invoice Total', 'Sum', 'Value (INR)', 'Sales Amount',
                  'Purchase Amount', 'Purchase Value']
CUST_NAME_KEYS = ['Bill-to Name', 'Sell-to Customer Name', 'Customer Name',
                  'Bill To Name', 'Bill To', 'Buyer Name', 'Buyer',
                  'Customer', 'Client Name', 'Client', 'Sold To',
                  'Sold-to', 'Sold-to Customer Name', 'Receiver',
                  'Receiver Name', 'Trade Name', 'Firm Name', 'Account Holder',
                  'Party Name', 'Account Name']
COUNTRY_KEYS   = ['Ship-to Country/Region Code', 'Sell-to Country/Region Code',
                  'Country/Region Code', 'Country', 'Country Code',
                  'Place of Supply', 'Customer State', 'State Name',
                  'State Code', 'State']
CURRENCY_KEYS  = ['Currency Code', 'Currency']

sales_hdr = []
sales_lines = []
for p in all_files('sales'):
    hdr = load_named_sheet(p, 'Sales Invoice Header')
    lns = load_named_sheet(p, 'Sales Invoice Line')
    if not hdr and not lns:
        # Flat sales register — every row is both header AND line. Customer
        # name + amount sit on the same row, no join required.
        flat = load_sheet(p)
        if not flat:
            flat = load_sheet(p, header_row=1)
        hdr = flat
        lns = flat
    sales_hdr.extend(hdr)
    sales_lines.extend(lns)
print(f"  Sales: headers={len(sales_hdr)} lines={len(sales_lines)}")

# Build doc_no -> header lookup using fuzzy doc-no detection.
hdr_by_doc = {}
for r in sales_hdr:
    k = find_value(r, DOC_NO_KEYS)
    if k is not None:
        hdr_by_doc[k] = r

# Top customers: derive name from the header row if the join works, otherwise
# fall back to the line row itself (flat-register case where the line *is*
# the header). Rows that resolve to neither a name nor a non-zero amount are
# dropped so we don't end up with a "Unidentified : 0" entry on the deck.
cust_rev = defaultdict(float)
cust_count = defaultdict(int)
country_rev = defaultdict(float)
currency_rev = defaultdict(float)
for ln in sales_lines:
    doc = find_value(ln, DOC_NO_KEYS)
    amt = float(find_value(ln, AMOUNT_KEYS) or 0)
    h = hdr_by_doc.get(doc) or ln
    raw_name = find_value(h, CUST_NAME_KEYS)
    name = str(raw_name).strip() if raw_name else 'Unidentified'
    if name == 'Unidentified' and amt == 0:
        continue  # no signal — skip silently
    cust_rev[name] += amt
    cust_count[name] += 1
    country_rev[find_value(h, COUNTRY_KEYS) or 'IN'] += amt
    currency_rev[find_value(h, CURRENCY_KEYS) or 'INR'] += amt

# Strip the all-zero "Unidentified" bucket if it slipped through.
if cust_rev.get('Unidentified', 0) == 0:
    cust_rev.pop('Unidentified', None)

top_customers = sorted(cust_rev.items(), key=lambda x: -x[1])[:10]
print(f"\nTop 10 customers by invoice line amount:")
for n, a in top_customers:
    print(f"  {n[:40]:<40}: {a:>14,.0f}")

# Diagnostic: if extraction produced nothing useful, dump column headers so
# the next iteration can extend the candidate list to whatever the user's
# file actually uses. Visible in the Trigger.dev run trace.
if not top_customers or sum(a for _, a in top_customers) == 0:
    print(f"[diag] top_customers empty — sales_lines headers seen: {list_headers(sales_lines)}")
    print(f"[diag] top_customers empty — sales_hdr  headers seen: {list_headers(sales_hdr)}")

unique_customer_count = len(cust_rev)

# Items — uploaded by the user as "inventory" (Item Ledger Entry export).
# Optional: if no inventory file was provided, top_items just stays empty
# instead of crashing.
print("\nLoading item ledger...")
items = []
for p in all_files('inventory'):
    rows = load_sheet(p)
    if not rows:
        rows = load_sheet(p, header_row=1)
    print(f"  + {os.path.basename(p)}: {len(rows)} rows")
    items.extend(rows)
items_sold = defaultdict(lambda: {'qty': 0, 'count': 0})
for r in items:
    if r.get('Entry Type') == 'Sale':
        qty = abs(float(r.get('Quantity') or 0))
        items_sold[r.get('Item No.')]['qty'] += qty
        items_sold[r.get('Item No.')]['count'] += 1

top_items = sorted(items_sold.items(), key=lambda x: -x[1]['qty'])[:10]

# Vendors — same fuzzy-column treatment as sales. Top vendors come from the
# purchase data, NOT vendor aging (aging is used downstream for DPO / cash
# flow only).
VEND_NAME_KEYS = ['Pay-to Name', 'Buy-from Vendor Name', 'Vendor Name',
                  'Supplier Name', 'Vendor', 'Supplier', 'Seller',
                  'Seller Name', 'Bill From', 'From', 'Vendor Code',
                  'Supplier Code', 'Buy-from Vendor No.', 'Trade Name',
                  'Firm Name', 'Party Name', 'Account Name']

pur_hdr = []
pur_lines = []
for p in all_files('purchase'):
    hdr = load_named_sheet(p, 'Purch. Inv. Header')
    lns = load_named_sheet(p, 'Purch. Inv. Line')
    if not hdr and not lns:
        flat = load_sheet(p)
        if not flat:
            flat = load_sheet(p, header_row=1)
        hdr = flat
        lns = flat
    pur_hdr.extend(hdr)
    pur_lines.extend(lns)
print(f"  Purchase: headers={len(pur_hdr)} lines={len(pur_lines)}")

phdr_by_doc = {}
for r in pur_hdr:
    k = find_value(r, DOC_NO_KEYS)
    if k is not None:
        phdr_by_doc[k] = r

vend_pur = defaultdict(float)
for ln in pur_lines:
    doc = find_value(ln, DOC_NO_KEYS)
    amt = float(find_value(ln, AMOUNT_KEYS) or 0)
    h = phdr_by_doc.get(doc) or ln
    raw_name = find_value(h, VEND_NAME_KEYS)
    name = str(raw_name).strip() if raw_name else 'Unidentified Vendor'
    if name == 'Unidentified Vendor' and amt == 0:
        continue
    vend_pur[name] += amt

if vend_pur.get('Unidentified Vendor', 0) == 0:
    vend_pur.pop('Unidentified Vendor', None)

top_vendors = sorted(vend_pur.items(), key=lambda x: -x[1])[:10]
print(f"\nTop 10 vendors by purchase line amount:")
for n, a in top_vendors:
    print(f"  {n[:40]:<40}: {a:>14,.0f}")

if not top_vendors or sum(a for _, a in top_vendors) == 0:
    print(f"[diag] top_vendors empty — pur_lines headers seen: {list_headers(pur_lines)}")
    print(f"[diag] top_vendors empty — pur_hdr   headers seen: {list_headers(pur_hdr)}")

unique_vendor_count = len(vend_pur)

# AR / AP - sourced from GL closing balances (account ranges)
# AR = 2310 (Trade Debtors) + 2320 (other receivables)
ar_balance = 0
for fy in ('FY23-24', 'FY24-25'):
    for acct in ('2310', '2320', '2330'):
        ar_balance += by_fy_acct[fy].get(acct, 0)

# AP = 5410 + 5420 (trade creditors)
ap_balance = 0
for fy in ('FY23-24', 'FY24-25'):
    for acct in ('5410', '5420'):
        ap_balance += by_fy_acct[fy].get(acct, 0)
ap_balance = -ap_balance  # liability accounts are negative balances; flip

total_ar = ar_balance
total_ap = ap_balance

# Per-customer / per-vendor open balances unavailable from ledger exports
ar_open = {}
ap_open = {}

# Monthly revenue trend FY24-25
monthly_rev = {}
for ym, cats in sorted(by_month_cat.items()):
    if ym >= '2024-04' and ym <= '2025-02':
        monthly_rev[ym] = -cats.get('Revenue', 0)

print(f"\nMonthly Revenue trend (FY24-25):")
for ym, r in monthly_rev.items():
    print(f"  {ym}: {r:>14,.0f}")

# ---------- Budget vs Actual variance ----------
import re

INCOME_KEYS = [
    (re.compile(r"\brev|sales\b|turnover|top[\s-]?line", re.I), 'Revenue'),
    (re.compile(r"gross\s*profit", re.I), 'Gross Profit'),
    (re.compile(r"other\s*income", re.I), 'Other Income'),
]
EXPENSE_KEYS = [
    (re.compile(r"\bcogs\b|cost\s*of\s*(goods|sales|sold)", re.I), 'COGS'),
    (re.compile(r"operat(ing|ions)\s*(exp|cost)|opex", re.I), 'Operating Expenses'),
    (re.compile(r"depreciat|amortis", re.I), 'Depreciation'),
    (re.compile(r"finance|interest", re.I), 'Finance Costs'),
    (re.compile(r"other\s*exp", re.I), 'Other Expenses'),
    (re.compile(r"\btax\b", re.I), 'Tax (Indicative @25.17%)'),
]
SUBTOTAL_KEYS = [
    (re.compile(r"ebitda", re.I), 'EBITDA'),
    (re.compile(r"\bebit\b(?!da)", re.I), 'EBIT'),
    (re.compile(r"\bpbt\b|profit\s*before\s*tax", re.I), 'PBT'),
    (re.compile(r"\bpat\b|profit\s*after\s*tax|net\s*profit", re.I), 'PAT'),
]


def _pick_budget_cols(rows):
    if not rows:
        return None, None, None
    headers = [h for h in rows[0].keys() if h]
    item_col = None; budget_col = None; cat_col = None
    for h in headers:
        lh = str(h).lower().strip()
        if not item_col and any(t in lh for t in ('line item','particular','account','head','description','name','item')):
            item_col = h
        if not budget_col and ('budget' in lh or 'plan' in lh or 'target' in lh) and any(t in lh for t in ('amount','amt','value','inr','rs','cr','lakh','fy')):
            budget_col = h
        if not budget_col and lh == 'budget':
            budget_col = h
        if not cat_col and any(t in lh for t in ('type','category','nature','kind','class')):
            cat_col = h
    if not budget_col:
        # first numeric-ish column
        for h in headers:
            for r in rows[:10]:
                v = r.get(h)
                if isinstance(v, (int, float)) and v != 0:
                    budget_col = h; break
                if isinstance(v, str) and re.search(r'\d', v):
                    budget_col = h; break
            if budget_col: break
    if not item_col:
        for h in headers:
            if h != budget_col:
                item_col = h; break
    return item_col, budget_col, cat_col


def _to_num(v):
    if isinstance(v, (int, float)):
        return float(v)
    if not isinstance(v, str):
        return 0.0
    s = re.sub(r'[^0-9.()-]', '', v)
    neg = s.startswith('(')
    s = s.replace('(', '').replace(')', '')
    try:
        n = float(s)
    except Exception:
        return 0.0
    return -n if neg else n


variance_rows = []
budget_paths = all_files('budget')
if budget_paths:
    print(f"\nLoading Budget file(s): {len(budget_paths)}")
    bud_rows = []
    for p in budget_paths:
        rs = load_sheet(p)
        if not rs:
            rs = load_sheet(p, header_row=1)
        bud_rows.extend(rs)
        print(f"  + {os.path.basename(p)}: {len(rs)} rows")
    item_col, budget_col, cat_col = _pick_budget_cols(bud_rows)
    if item_col and budget_col:
        by_item = {}
        for r in bud_rows:
            name = str(r.get(item_col) or '').strip()
            if not name:
                continue
            amt = _to_num(r.get(budget_col))
            if not amt:
                continue
            cat = str(r.get(cat_col) or '').lower() if cat_col else ''
            kind = None
            if any(t in cat for t in ('income','revenue','sales','gain')):
                kind = 'income'
            elif any(t in cat for t in ('expen','cost','cogs','tax')):
                kind = 'expense'
            prev = by_item.get(name, {'amount': 0.0, 'kind': kind})
            prev['amount'] += amt
            if not prev.get('kind') and kind:
                prev['kind'] = kind
            by_item[name] = prev

        for name, rec in by_item.items():
            pl_key = None
            kind = rec.get('kind') or 'expense'
            for rx, k in INCOME_KEYS:
                if rx.search(name):
                    pl_key = k; kind = 'income'; break
            if not pl_key:
                for rx, k in EXPENSE_KEYS:
                    if rx.search(name):
                        pl_key = k; kind = 'expense'; break
            if not pl_key:
                for rx, k in SUBTOTAL_KEYS:
                    if rx.search(name):
                        pl_key = k; kind = 'income'; break

            actual = pl_24_25_ann.get(pl_key, 0.0) if pl_key else 0.0
            budget_amt = rec['amount']
            var = actual - budget_amt
            var_pct = (var / abs(budget_amt) * 100) if budget_amt else 0.0
            favorable = (var >= 0) if kind == 'income' else (var <= 0)
            variance_rows.append({
                'lineItem': name,
                'kind': kind,
                'budget': budget_amt,
                'actual': actual,
                'variance': var,
                'variancePct': var_pct,
                'favorable': favorable,
            })
        # Sort income first, then by abs variance desc
        variance_rows.sort(key=lambda r: (0 if r['kind'] == 'income' else 1, -abs(r['variance'])))
        print(f"  Variance rows: {len(variance_rows)}")
    else:
        print("  Could not detect budget columns — skipping variance analysis")


# ---------- Customer / Vendor Aging + Cash Flow ----------
from datetime import datetime as _dt

AGING_BUCKETS = ['Current', '0-30', '31-60', '61-90', '91-180', '180+']

_BUCKET_PATTERNS = [
    (re.compile(r"^current$|^not\s*due$", re.I), 'Current'),
    (re.compile(r"^0[\s-]*30$|^<=?\s*30$|^upto\s*30", re.I), '0-30'),
    (re.compile(r"^31[\s-]*60$|^>30.*<=?60", re.I), '31-60'),
    (re.compile(r"^61[\s-]*90$|^>60.*<=?90", re.I), '61-90'),
    (re.compile(r"^91[\s-]*180$|^>90.*<=?180", re.I), '91-180'),
    (re.compile(r"^180\+$|^>180|^over\s*180", re.I), '180+'),
]


def _bucketize(days):
    if days <= 0: return 'Current'
    if days <= 30: return '0-30'
    if days <= 60: return '31-60'
    if days <= 90: return '61-90'
    if days <= 180: return '91-180'
    return '180+'


def _find_col(headers, regex_list):
    for h in headers:
        if not h:
            continue
        lh = str(h).lower().strip()
        for rx in regex_list:
            if rx.search(lh):
                return h
    return None


def parse_aging_files(paths, as_of):
    if not paths:
        return None
    all_rows = []
    for p in paths:
        rs = load_sheet(p)
        if not rs:
            rs = load_sheet(p, header_row=1)
        all_rows.extend(rs)
        print(f"  + aging file {os.path.basename(p)}: {len(rs)} rows")
    if not all_rows:
        return None

    headers = list((all_rows[0] or {}).keys())
    party_col = _find_col(headers, [re.compile(r"customer|vendor|party|account|name|debtor|creditor|supplier", re.I)])
    if not party_col:
        print("  Could not detect party column — skipping")
        return None

    bucket_cols = []
    for h in headers:
        norm = str(h).lower().strip().replace(' ', '')
        for pat, bk in _BUCKET_PATTERNS:
            if pat.search(norm):
                bucket_cols.append((h, bk))
                break

    buckets = {b: 0.0 for b in AGING_BUCKETS}
    party_totals = {}  # name -> {amount, topBucket, daysOverdue?}
    total = 0.0

    if len(bucket_cols) >= 2:
        # Shape A
        for r in all_rows:
            name = str(r.get(party_col) or '').strip()
            if not name:
                continue
            party_amt = 0.0
            worst = 'Current'
            for col, bk in bucket_cols:
                v = _to_num(r.get(col))
                if v:
                    buckets[bk] += v
                    party_amt += v
                    total += v
                    if AGING_BUCKETS.index(bk) > AGING_BUCKETS.index(worst):
                        worst = bk
            if party_amt:
                if name in party_totals:
                    party_totals[name]['amount'] += party_amt
                    if AGING_BUCKETS.index(worst) > AGING_BUCKETS.index(party_totals[name]['topBucket']):
                        party_totals[name]['topBucket'] = worst
                else:
                    party_totals[name] = {'amount': party_amt, 'topBucket': worst}
    else:
        # Shape B - invoice-level dates
        amt_col = _find_col(headers, [re.compile(r"amount|outstanding|balance|due|net", re.I)])
        due_col = _find_col(headers, [re.compile(r"due\s*date", re.I)])
        inv_col = _find_col(headers, [re.compile(r"invoice\s*date|bill\s*date|posting\s*date|document\s*date", re.I)])
        if not amt_col:
            print("  Could not detect amount column — skipping")
            return None
        for r in all_rows:
            name = str(r.get(party_col) or '').strip()
            if not name:
                continue
            amt = _to_num(r.get(amt_col))
            if not amt:
                continue
            ref = to_date(r.get(due_col)) if due_col else None
            if ref is None and inv_col:
                ref = to_date(r.get(inv_col))
            days = (as_of - ref).days if ref else 0
            bk = _bucketize(days)
            buckets[bk] += amt
            total += amt
            if name in party_totals:
                party_totals[name]['amount'] += amt
                if AGING_BUCKETS.index(bk) > AGING_BUCKETS.index(party_totals[name]['topBucket']):
                    party_totals[name]['topBucket'] = bk
                    party_totals[name]['daysOverdue'] = days
            else:
                party_totals[name] = {'amount': amt, 'topBucket': bk, 'daysOverdue': days}

    top_parties = sorted(
        [{'name': n, 'amount': v['amount'], 'bucket': v['topBucket'], 'daysOverdue': v.get('daysOverdue')}
         for n, v in party_totals.items()],
        key=lambda x: -x['amount']
    )[:10]

    return {
        'asOfDate': as_of.isoformat() if hasattr(as_of, 'isoformat') else str(as_of),
        'totalOutstanding': total,
        'buckets': buckets,
        'topParties': top_parties,
        'partyCount': len(party_totals),
    }


_as_of = latest_posting_date or _dt.now().date()
customer_aging = None
vendor_aging = None
cash_flow = None

_ca_paths = all_files('customerAging')
if _ca_paths:
    print(f"\nLoading Customer Aging files: {len(_ca_paths)}")
    customer_aging = parse_aging_files(_ca_paths, _as_of)

_va_paths = all_files('vendorAging')
if _va_paths:
    print(f"\nLoading Vendor Aging files: {len(_va_paths)}")
    vendor_aging = parse_aging_files(_va_paths, _as_of)

if customer_aging or vendor_aging:
    def _collect(a):
        if not a: return {'0-30': 0.0, '31-60': 0.0, '61-90': 0.0}
        b = a['buckets']
        return {
            '0-30':  b.get('Current', 0) + b.get('0-30', 0) * 0.9,
            '31-60': b.get('31-60', 0) * 0.7,
            '61-90': b.get('61-90', 0) * 0.5,
        }
    def _pay(a):
        if not a: return {'0-30': 0.0, '31-60': 0.0, '61-90': 0.0}
        b = a['buckets']
        return {
            '0-30':  b.get('Current', 0) + b.get('0-30', 0),
            '31-60': b.get('31-60', 0),
            '61-90': b.get('61-90', 0),
        }
    c = _collect(customer_aging)
    p = _pay(vendor_aging)
    windows = []
    for label, key in [('Next 30 days', '0-30'), ('31-60 days', '31-60'), ('61-90 days', '61-90')]:
        windows.append({
            'label': label,
            'collections': c[key],
            'payments': p[key],
            'net': c[key] - p[key],
        })
    total_c = sum(w['collections'] for w in windows)
    total_p = sum(w['payments'] for w in windows)
    cash_flow = {
        'windows': windows,
        'totalCollections': total_c,
        'totalPayments': total_p,
        'netCashFlow': total_c - total_p,
    }
    print(f"  Cash Flow: collections {total_c:,.0f}  payments {total_p:,.0f}  net {total_c - total_p:,.0f}")


# ========================================================================
# Analytics layer — mirrors web/lib/financials.ts engine
# ========================================================================
BENCH = {
    'gross_margin_good': 35, 'gross_margin_poor': 25,
    'ebitda_margin_good': 10, 'ebitda_margin_poor': 5,
    'pat_margin_good': 5,
    'dso_good': 45, 'dso_poor': 75,
    'dpo_good': 45, 'dpo_poor': 90,
    'current_ratio_good': 1.5, 'current_ratio_poor': 1.0,
    'customer_concentration_warn': 30,
    'opex_line_warn': 5,
}


def _fmt_cr(v):
    cr = v / 1e7
    sign = '−' if cr < 0 else ''
    return f"{sign}₹{abs(cr):.2f} Cr"


def _fmt_pct(v):
    sign = '−' if v < 0 else ''
    return f"{sign}{abs(v):.1f}%"


def _in_range(totals, lo, hi):
    s = 0.0
    for acct, v in totals.items():
        try:
            a = int(str(acct))
        except Exception:
            continue
        if lo <= a <= hi:
            s += v
    return s


# Build cross-FY totals
_acct_totals_all = {}
for acct_map in by_fy_acct.values():
    for acct, v in acct_map.items():
        _acct_totals_all[acct] = _acct_totals_all.get(acct, 0) + v

# Balance Sheet
balance_sheet = None
if _acct_totals_all:
    fixed = _in_range(_acct_totals_all, 1000, 1999)
    inv = _in_range(_acct_totals_all, 2000, 2199)
    rec = _in_range(_acct_totals_all, 2200, 2399)
    other_a = _in_range(_acct_totals_all, 2400, 2799)
    cash = _in_range(_acct_totals_all, 2800, 2999)
    total_assets = fixed + inv + rec + other_a + cash

    equity = -_in_range(_acct_totals_all, 3000, 3999)
    lt_debt = -_in_range(_acct_totals_all, 4000, 4999)
    payables = -_in_range(_acct_totals_all, 5000, 5499)
    other_cl = -_in_range(_acct_totals_all, 5500, 5999)
    total_le = equity + lt_debt + payables + other_cl

    cur_assets = inv + rec + other_a + cash
    cur_liab = payables + other_cl
    current_ratio = (cur_assets / cur_liab) if cur_liab else 0
    debt_to_equity = (lt_debt / equity) if equity else 0

    balance_sheet = {
        'fixedAssetsNet': fixed, 'inventory': inv, 'receivables': rec,
        'cashAndBank': cash, 'otherAssets': other_a, 'totalAssets': total_assets,
        'equity': equity, 'longTermDebt': lt_debt, 'payables': payables,
        'otherCurrentLiab': other_cl, 'totalLiabilitiesAndEquity': total_le,
        'currentRatio': current_ratio, 'debtToEquity': debt_to_equity,
    }

# Cost Structure (target FY)
cost_structure = []
_target_acct_map = by_fy_acct.get('FY24-25', {})
_cost_buckets = [
    ("Raw Material / Direct Cost",   7000, 7299),
    ("Other Direct Costs",           7300, 7999),
    ("Salaries & Personnel",         8100, 8199),
    ("Office & Admin",               8200, 8299),
    ("Sales & Marketing",            8300, 8399),
    ("Legal & Professional",         8400, 8499),
    ("Utilities & Communications",   8500, 8599),
    ("Travel & Conveyance",          8600, 8699),
    ("Other Operating Expenses",     8700, 8799),
    ("Depreciation & Amortization",  8800, 8899),
    ("Miscellaneous OpEx",           8900, 8999),
    ("Finance Costs",                9500, 9999),
]
_rev_actual = pl_24_25['Revenue']
for name, lo, hi in _cost_buckets:
    amt = _in_range(_target_acct_map, lo, hi)
    if abs(amt) < 1:
        continue
    pct = (amt / _rev_actual * 100) if _rev_actual else 0
    cost_structure.append({
        'category': name,
        'amount': amt,
        'percentOfRevenue': pct,
        'isWatchlist': pct > BENCH['opex_line_warn'],
        'benchmark': f"Above {BENCH['opex_line_warn']}% threshold — review" if pct > BENCH['opex_line_warn'] else None,
    })
cost_structure.sort(key=lambda x: -x['amount'])

# Benchmarks
benchmarks = []
gm = pl_24_25['Gross Margin %']
benchmarks.append({
    'metric': 'Gross Margin',
    'actual': _fmt_pct(gm), 'actualNumeric': gm,
    'benchmark': f"{BENCH['gross_margin_poor']}–{BENCH['gross_margin_good']}%",
    'status': 'good' if gm >= BENCH['gross_margin_good'] else ('ok' if gm >= BENCH['gross_margin_poor'] else 'poor'),
    'gap': f"{(BENCH['gross_margin_good'] - gm):.1f}pp below target" if gm < BENCH['gross_margin_good'] else "At or above target",
})
em = pl_24_25['EBITDA Margin %']
benchmarks.append({
    'metric': 'EBITDA Margin',
    'actual': _fmt_pct(em), 'actualNumeric': em,
    'benchmark': f"{BENCH['ebitda_margin_poor']}–{BENCH['ebitda_margin_good']}%",
    'status': 'good' if em >= BENCH['ebitda_margin_good'] else ('ok' if em >= BENCH['ebitda_margin_poor'] else 'poor'),
    'gap': f"{(BENCH['ebitda_margin_good'] - em):.1f}pp below target" if em < BENCH['ebitda_margin_good'] else "At or above target",
})
nm = pl_24_25['PAT / Net Margin %']
benchmarks.append({
    'metric': 'Net Margin (PAT)',
    'actual': _fmt_pct(nm), 'actualNumeric': nm,
    'benchmark': f"≥ {BENCH['pat_margin_good']}%",
    'status': 'good' if nm >= BENCH['pat_margin_good'] else ('ok' if nm >= 0 else 'poor'),
    'gap': "Company operating at a loss" if nm < 0 else (f"{(BENCH['pat_margin_good'] - nm):.1f}pp below target" if nm < BENCH['pat_margin_good'] else "At or above target"),
})
# DSO / DPO using balance sheet
if balance_sheet and _rev_actual:
    dso = (balance_sheet['receivables'] / _rev_actual) * 365
    benchmarks.append({
        'metric': 'DSO (Days Sales Outstanding)',
        'actual': f"{dso:.0f} days", 'actualNumeric': dso,
        'benchmark': f"{BENCH['dso_good']}–{BENCH['dso_poor']} days",
        'status': 'good' if dso <= BENCH['dso_good'] else ('ok' if dso <= BENCH['dso_poor'] else 'poor'),
        'gap': f"{(dso - BENCH['dso_good']):.0f} days above target" if dso > BENCH['dso_good'] else "Within target",
    })
if balance_sheet and pl_24_25['COGS']:
    dpo = (balance_sheet['payables'] / pl_24_25_ann['COGS']) * 365
    benchmarks.append({
        'metric': 'DPO (Days Payables Outstanding)',
        'actual': f"{dpo:.0f} days", 'actualNumeric': dpo,
        'benchmark': f"{BENCH['dpo_good']}–{BENCH['dpo_poor']} days",
        'status': 'good' if BENCH['dpo_good'] <= dpo <= BENCH['dpo_poor'] else ('ok' if dpo < BENCH['dpo_good'] else 'poor'),
        'gap': f"{(dpo - BENCH['dpo_poor']):.0f} days above range" if dpo > BENCH['dpo_poor'] else (f"{(BENCH['dpo_good'] - dpo):.0f} days below range" if dpo < BENCH['dpo_good'] else "Within range"),
    })
if balance_sheet and balance_sheet.get('currentRatio'):
    cr = balance_sheet['currentRatio']
    benchmarks.append({
        'metric': 'Current Ratio',
        'actual': f"{cr:.2f}x", 'actualNumeric': cr,
        'benchmark': f"≥ {BENCH['current_ratio_good']}x",
        'status': 'good' if cr >= BENCH['current_ratio_good'] else ('ok' if cr >= BENCH['current_ratio_poor'] else 'poor'),
        'gap': f"{(BENCH['current_ratio_good'] - cr):.2f}x below target — limited liquidity buffer" if cr < BENCH['current_ratio_good'] else "Adequate liquidity",
    })

# Critical Issues
critical_issues = []
if pl_24_25['EBIT'] < 0:
    critical_issues.append({
        'title': 'Operating Loss — Costs Exceed Operating Income',
        'rootCause': f"EBIT is negative at {_fmt_cr(pl_24_25['EBIT'])} ({_fmt_pct(pl_24_25['EBIT Margin %'])} margin). OpEx ({_fmt_cr(pl_24_25['Operating Expenses'])}) plus depreciation ({_fmt_cr(pl_24_25['Depreciation'])}) outweigh gross profit.",
        'recommendedAction': "Cut top operating expense lines; renegotiate supplier contracts; review fixed-cost base.",
        'potentialImpact': f"Closing the EBIT gap would unlock {_fmt_cr(abs(pl_24_25['EBIT']))} in annual operating profit.",
        'severity': 'high',
    })
if pl_24_25['EBITDA'] > 0 and em < BENCH['ebitda_margin_good']:
    critical_issues.append({
        'title': 'EBITDA Margin Below Industry Benchmark',
        'rootCause': f"EBITDA margin at {_fmt_pct(em)} vs industry target of {BENCH['ebitda_margin_good']}%+. Buffer for cost shocks is thin.",
        'recommendedAction': "Identify top 3 OpEx lines as % of revenue and target 10-20% reduction in each.",
        'potentialImpact': f"Each 1pp of EBITDA margin = {_fmt_cr(_rev_actual * 0.01)} annual profit.",
        'severity': 'high',
    })
if gm < BENCH['gross_margin_good']:
    critical_issues.append({
        'title': f"Gross Margin at {_fmt_pct(gm)} — Below {BENCH['gross_margin_good']}% Target",
        'rootCause': f"Gross margin gap of {(BENCH['gross_margin_good'] - gm):.1f}pp. COGS at {_fmt_pct((pl_24_25['COGS']/_rev_actual*100) if _rev_actual else 0)} of revenue.",
        'recommendedAction': "Review pricing on top SKUs; consolidate raw material vendors for volume discount; pass-through input cost increases.",
        'potentialImpact': f"Each 1pp of gross margin gain = {_fmt_cr(_rev_actual * 0.01)} additional gross profit.",
        'severity': 'high' if gm < BENCH['gross_margin_poor'] else 'medium',
    })
_dso_b = next((b for b in benchmarks if b['metric'].startswith('DSO')), None)
if _dso_b and _dso_b['actualNumeric'] > BENCH['dso_poor']:
    cash_locked = (_dso_b['actualNumeric'] - BENCH['dso_good']) / 365 * _rev_actual
    critical_issues.append({
        'title': f"Slow Collections — DSO at {_dso_b['actual']}",
        'rootCause': f"Receivables conversion is taking {_dso_b['actual']} vs target of {BENCH['dso_good']} days.",
        'recommendedAction': "Tighten credit terms; offer early-payment discounts; assign dedicated collections owner for accounts > 60 days.",
        'potentialImpact': f"Reducing DSO to {BENCH['dso_good']} days would release {_fmt_cr(cash_locked)} in working capital.",
        'severity': 'high',
    })
if top_customers:
    total_top = sum(a for _, a in top_customers)
    if total_top > 0:
        top_share = (top_customers[0][1] / total_top) * 100
        if top_share > BENCH['customer_concentration_warn']:
            critical_issues.append({
                'title': 'Customer Concentration Risk',
                'rootCause': f'Top customer "{top_customers[0][0]}" accounts for {top_share:.1f}% of sampled invoice value.',
                'recommendedAction': "Diversify customer base; cap any single customer at < 20% of revenue.",
                'potentialImpact': f"Reducing concentration de-risks {_fmt_cr(top_customers[0][1])} of revenue exposure.",
                'severity': 'high' if top_share > 50 else 'medium',
            })
if balance_sheet and balance_sheet.get('currentRatio') and balance_sheet['currentRatio'] < BENCH['current_ratio_good']:
    cr = balance_sheet['currentRatio']
    critical_issues.append({
        'title': f"Current Ratio at {cr:.2f}x — Liquidity Pressure",
        'rootCause': f"Current ratio below {BENCH['current_ratio_good']}x benchmark.",
        'recommendedAction': "Secure working capital line; extend supplier terms; accelerate receivables.",
        'potentialImpact': f"Bringing ratio to {BENCH['current_ratio_good']}x requires net working capital improvement.",
        'severity': 'high' if cr < BENCH['current_ratio_poor'] else 'medium',
    })

sev_order = {'high': 0, 'medium': 1, 'low': 2, 'info': 3}
critical_issues.sort(key=lambda x: sev_order.get(x['severity'], 9))
for i, issue in enumerate(critical_issues):
    issue['rank'] = i + 1
critical_issues = critical_issues[:5]

# Growth Opportunities
growth_opportunities = []
if len(top_customers) >= 3:
    top3 = sum(a for _, a in top_customers[:3])
    growth_opportunities.append({
        'rank': 1,
        'title': "Replicate Top Customer Profile",
        'rationale': f"Top 3 customers ({', '.join(n for n, _ in top_customers[:3])}) generated {_fmt_cr(top3)}.",
        'approach': "Map common attributes (industry, size, region) and run targeted outreach to look-alikes.",
        'potentialUpside': f"Each additional similar account = {_fmt_cr(top3 / 3)} potential revenue.",
    })

if country_rev:
    country_list = sorted(country_rev.items(), key=lambda x: -x[1])
    total_cr = sum(country_rev.values())
    if total_cr and len(country_list) >= 2:
        dom = country_list[0]
        share = (dom[1] / total_cr) * 100
        if share < 70:
            growth_opportunities.append({
                'rank': 2,
                'title': "Geographic Expansion",
                'rationale': f"Revenue is spread across {len(country_list)} countries with no single market dominating (top market {dom[0]} at {share:.0f}%).",
                'approach': f"Double down on the next 2-3 markets ({', '.join(c for c, _ in country_list[1:4])}).",
                'potentialUpside': "Doubling revenue in #2-3 markets could add meaningful incremental revenue.",
            })

if monthly_rev and len(monthly_rev) >= 6:
    items_list = sorted(monthly_rev.items())
    first = items_list[0][1]
    last = next((v for _, v in reversed(items_list) if v > 1000), first)
    growth = ((last / first) - 1) * 100 if first else 0
    if growth > 30:
        growth_opportunities.append({
            'rank': 3,
            'title': "Lock In Revenue Momentum",
            'rationale': f"Monthly revenue grew {growth:.0f}% from start to latest month. Strong demand signal.",
            'approach': "Invest in production capacity; sign long-term contracts; pre-empt seasonality.",
            'potentialUpside': f"Sustaining {growth:.0f}% growth for another quarter = {_fmt_cr(last * 3 * (growth / 100))} additional revenue.",
        })

growth_opportunities = growth_opportunities[:4]

# Insights
insights = []
if monthly_rev and len(monthly_rev) >= 3:
    items_list = sorted(monthly_rev.items())
    first = items_list[0][1]
    last = next((v for _, v in reversed(items_list) if v > 1000), first)
    growth = ((last / first) - 1) * 100 if first else 0
    insights.append({
        'category': 'revenue',
        'title': f"Revenue {'grew' if growth >= 0 else 'declined'} {abs(growth):.0f}% over the period",
        'detail': f"Monthly run-rate moved from {_fmt_cr(first)} to {_fmt_cr(last)}.",
        'severity': 'info' if growth > 30 else ('low' if growth > 0 else 'medium'),
    })

insights.append({
    'category': 'margin',
    'title': f"Gross margin at {_fmt_pct(gm)}",
    'detail': (
        f"Above industry target ({BENCH['gross_margin_good']}%+) — pricing and sourcing are working."
        if gm >= BENCH['gross_margin_good']
        else (
            f"Within acceptable range but {(BENCH['gross_margin_good'] - gm):.1f}pp below best-in-class."
            if gm >= BENCH['gross_margin_poor']
            else f"Below {BENCH['gross_margin_poor']}% — pricing power and input cost discipline both need work."
        )
    ),
    'severity': 'info' if gm >= BENCH['gross_margin_good'] else ('low' if gm >= BENCH['gross_margin_poor'] else 'high'),
})

insights.append({
    'category': 'margin',
    'title': f"EBITDA {'positive' if pl_24_25['EBITDA'] >= 0 else 'negative'} at {_fmt_cr(pl_24_25['EBITDA'])}",
    'detail': (
        "Operating costs exceed gross profit. Top OpEx lines need immediate attention."
        if pl_24_25['EBITDA'] < 0
        else (
            f"Margin of {_fmt_pct(em)} is below the {BENCH['ebitda_margin_good']}% benchmark. Limited cushion for cost shocks."
            if em < BENCH['ebitda_margin_good']
            else f"Margin of {_fmt_pct(em)} is healthy — sustainable cash generation."
        )
    ),
    'severity': 'high' if pl_24_25['EBITDA'] < 0 else ('medium' if em < BENCH['ebitda_margin_good'] else 'info'),
})

if total_ar and total_ap:
    nwc = total_ar - total_ap
    insights.append({
        'category': 'cashflow',
        'title': f"Net working capital deployed: {_fmt_cr(nwc)}",
        'detail': (
            f"Receivables ({_fmt_cr(total_ar)}) exceed payables ({_fmt_cr(total_ap)}). Cash is funding the working capital cycle."
            if nwc > 0
            else f"Payables ({_fmt_cr(total_ap)}) exceed receivables ({_fmt_cr(total_ar)}) — suppliers are funding operations."
        ),
        'severity': 'medium' if abs(nwc) > _rev_actual * 0.2 else 'info',
    })

if pl_24_25['PAT'] < 0:
    insights.append({
        'category': 'margin',
        'title': f"Net loss of {_fmt_cr(abs(pl_24_25['PAT']))}",
        'detail': "Path to profitability requires either revenue growth, gross margin expansion, or OpEx reduction.",
        'severity': 'high',
    })

# Improvement initiatives
improvements = []
for w in [c for c in cost_structure if c['isWatchlist']][:3]:
    improvements.append({
        'action': f"Reduce {w['category']} by 20%",
        'savings': abs(w['amount']) * 0.2,
        'timeline': "3-6 months",
        'difficulty': "medium",
        'rationale': f"Currently {_fmt_pct(w['percentOfRevenue'])} of revenue — above {BENCH['opex_line_warn']}% threshold.",
    })

if _dso_b and _rev_actual and _dso_b['actualNumeric'] > BENCH['dso_good']:
    days = min(15, _dso_b['actualNumeric'] - BENCH['dso_good'])
    improvements.append({
        'action': f"Reduce DSO by {days:.0f} days (collections discipline)",
        'savings': (days / 365) * _rev_actual,
        'timeline': "0-3 months",
        'difficulty': "easy",
        'rationale': "Stricter credit terms, early-payment discounts, dedicated collections owner.",
    })

if gm < BENCH['gross_margin_good']:
    pp_lift = min(3, BENCH['gross_margin_good'] - gm)
    improvements.append({
        'action': f"Lift gross margin by {pp_lift:.0f}pp via pricing + sourcing",
        'savings': (pp_lift / 100) * _rev_actual,
        'timeline': "6-12 months",
        'difficulty': "hard",
        'rationale': "Selective price increases on premium SKUs and consolidation of top-3 raw material vendors.",
    })
improvements.sort(key=lambda x: -x['savings'])

# AI Summary
ai_summary_parts = []
ai_summary_parts.append(
    f"The company posted revenue of {_fmt_cr(_rev_actual)} over the period (annualized: {_fmt_cr(pl_24_25_ann['Revenue'])}). "
    f"Gross margin came in at {_fmt_pct(gm)} and EBITDA at {_fmt_pct(em)}, producing "
    f"{'a net profit' if pl_24_25['PAT'] >= 0 else 'a net loss'} of {_fmt_cr(abs(pl_24_25['PAT']))} for the period."
)
if critical_issues:
    top_issue = critical_issues[0]
    ai_summary_parts.append(
        f"The most pressing issue is {top_issue['title'].lower()}. {top_issue['rootCause']} "
        f"Addressing the top {min(len(critical_issues), 3)} issues identified below could meaningfully improve the bottom line."
    )
else:
    ai_summary_parts.append("The financial profile is broadly healthy across margins, working capital and concentration metrics.")
if growth_opportunities:
    ai_summary_parts.append(
        f"On the upside, {len(growth_opportunities)} growth opportunities have been identified. "
        f"Execution discipline on the top opportunity ({growth_opportunities[0]['title'].lower()}) is likely the highest-ROI bet for the next 12 months."
    )
ai_summary = "\n\n".join(ai_summary_parts)


# ---------- Save JSON for downstream ----------
summary = {
    'period': 'FY 2024-25 (Apr 2024 - Feb 2025, 11 months actuals)',
    'pl_actual': pl_24_25,
    'pl_annualized': pl_24_25_ann,
    'pl_prior_partial': pl_23_24,
    'top_customers': top_customers,
    'top_items': [(k, v) for k,v in top_items],
    'top_vendors': top_vendors,
    'unique_customer_count': unique_customer_count,
    'unique_vendor_count': unique_vendor_count,
    'monthly_revenue': monthly_rev,
    'total_ar_open': total_ar,
    'total_ap_open': total_ap,
    'country_rev': dict(country_rev),
    'currency_rev': dict(currency_rev),
    'ar_open_top': [],
    'ap_open_top': [],
    'gl_account_summary': {fy: dict(d) for fy, d in by_fy_cat.items()},
    'variance': variance_rows,
    'has_budget': len(variance_rows) > 0,
    'customer_aging': customer_aging,
    'vendor_aging': vendor_aging,
    'cash_flow': cash_flow,
    'balance_sheet': balance_sheet,
    'cost_structure': cost_structure,
    'benchmarks': benchmarks,
    'critical_issues': critical_issues,
    'growth_opportunities': growth_opportunities,
    'insights': insights,
    'improvements': improvements,
    'ai_summary': ai_summary,
}

with open(os.path.join(TMP, 'financial_summary.json'), 'w') as f:
    json.dump(summary, f, indent=2, default=str)

print(f"\nSaved to {os.path.join(TMP, 'financial_summary.json')}")
print("\n=== Key Metrics for the deck ===")
print(f"FY24-25 Revenue (11M):   INR {pl_24_25['Revenue']:>15,.0f}")
print(f"FY24-25 EBITDA (11M):    INR {pl_24_25['EBITDA']:>15,.0f}")
print(f"FY24-25 PAT (11M):       INR {pl_24_25['PAT']:>15,.0f}")
print(f"Annualized Revenue:      INR {pl_24_25_ann['Revenue']:>15,.0f}")
print(f"Total open AR:           INR {total_ar:>15,.0f}")
print(f"Total open AP:           INR {total_ap:>15,.0f}")
